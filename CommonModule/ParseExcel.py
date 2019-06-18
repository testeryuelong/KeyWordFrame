# -*-coding:utf-8 -*-
# @Author : Zhigang

import os
from openpyxl import load_workbook
from Config.ProjectVar import *
from openpyxl.styles import Font
import traceback
from CommonModule.RiseTime import *

class ParseExcel(object):
    """封装常用excel方法"""

    def __init__(self,excelFilePath):
        if not os.path.exists(excelFilePath):
            self.excelFilePath=None
        else:
            self.excelFilePath=excelFilePath
        if self.excelFilePath:
            self.wb=load_workbook(self.excelFilePath)
            # 常用颜色代码
            self.RGBDict={"red": "00FF0000", "green": "0000FF00","black":"00000000","yellow":"00FFFF00"}

    def getExcelFilePath(self):
        "获取文件路径"
        return self.excelFilePath

    def getExcelAllSheetNames(self):
        "获取所有的sheet名称，结果为列表形式"
        return self.wb.sheetnames

    def getExcelAllSheetObj(self):
        "获取所有的sheet对象，结果为列表形式"
        sheetsObj=[]
        for sheetName in self.wb.sheetnames:
            sheetsObj.append(self.wb[sheetName])
        return sheetsObj

    def whetherExists(self,sheetName):
        "判断sheet名称是否在excel中，结果返回布尔值"
        if sheetName in self.getExcelAllSheetNames():
            return True
        else:
            return False

    def setExcelOprateSheet(self,sheetName):
        "设定要操作的sheet对象，结果返回sheet对象或None"
        if self.whetherExists(sheetName):
            self.sheet=self.wb[sheetName]
        else:
            self.sheet=None
        return self.sheet

    def getCurrentSheetName(self):
        "获取当前操作的sheet名字,结果返回字符串"
        return self.sheet.title

    def getMinRow(self):
        "获取当前操作的sheet最小行值,从1开始"
        return self.sheet.min_row

    def getMaxRow(self):
        "获取当前操作的sheet最大行值"
        return self.sheet.max_row

    def getMinColumn(self):
        "获取当前操作的sheet最小列值,从1开始"
        return self.sheet.min_column

    def getMaxColumn(self):
        "获取当前操作的sheet最大列值"
        return self.sheet.max_column

    def getTargetRowObj(self,rowNo):
        "获取某一行单元格对象"
        if not isinstance(rowNo,int):
            return "参数应为int类型"
        if rowNo >= self.getMinRow() and rowNo <= self.getMaxRow():
            return list(self.sheet.iter_rows())[rowNo-1]
        else:
            return "不在行号范围"

    def getTargetColumnObj(self,columnNo):
        "获取某一列单元格对象"
        if not isinstance(columnNo,int):
            return "参数应为int类型"
        if columnNo >= self.getMinColumn() and columnNo <= self.getMaxColumn():
            return list(self.sheet.iter_cols())[columnNo-1]
        else:
            return "不在列号范围"

    def getTargetCellObj(self,rowNo,columnNo):
        "获取指定单元格对象"
        if isinstance(rowNo,int) and isinstance(columnNo,int):
            return self.sheet.cell(rowNo,columnNo)
        else:
            return "参数应为int类型"

    def getTargetCellValue(self,rowNo,columnNo):
        "获取指定单元格对象"
        if isinstance(rowNo,int) and isinstance(columnNo,int):
            return self.sheet.cell(rowNo,columnNo).value
        else:
            return "参数应为int类型"

    def writeCellStrValue(self,rowNo,columnNo,content,style=None):
        "往指定的单元格写入内容并保存，颜色默认为黑色"
        try:
            targetCell=self.getTargetCellObj(rowNo,columnNo)
            targetCell.value=content
            if style == None:
                targetCell.font=Font(color=self.RGBDict["black"])
            else:
                targetCell.font=Font(color=self.RGBDict[style])
            self.wb.save(self.excelFilePath)
        except:
            traceback.print_exc()

    def clearColumnValue(self,sheetName,columnNo,startRow=2):
        "清空指定sheet的指定列号整列内容，默认从第二行开始清空，保留第一行字段内容"
        try:
            self.setExcelOprateSheet(sheetName)
            for rowNo in range(startRow,self.getMaxRow()+1):
                self.writeCellStrValue(rowNo,columnNo,content="")
        except:
            traceback.print_exc()

    def clearValue(self,sheetName,rowNo,columnNo):
        "清空指定sheet的指定单元格内容"
        try:
            self.setExcelOprateSheet(sheetName)
            self.writeCellStrValue(rowNo,columnNo,content="")
        except:
            traceback.print_exc()

    def clearSheetValue(self,sheetName):
        "如果测试用例需要执行，必定会清除下面四个列的值，故封装为函数"
        self.setExcelOprateSheet(sheetName)
        self.clearColumnValue(sheetName, testStepExecutedTime)
        self.clearColumnValue(sheetName, testStepExecutedResult)
        self.clearColumnValue(sheetName, testStepErrorInfo)
        self.clearColumnValue(sheetName, testStepScreenshotPath)

    def printTestMethod(self):
        # print (self.excelFilePath)
        # print (self.wb)
        # print (self.RGBDict["red"])
        # print (self.getExcelFilePath())
        # print (self.getExcelAllSheetNames())
        # print (self.whetherExists("百度"))
        # print(self.whetherExists("百度2"))
        # print (self.setExcelOprateSheet("百度2"))
        # print(self.setExcelOprateSheet("百度"))
        # print (self.getExcelAllSheetObj())
        # print (self.getCurrentSheetName())
        # print (self.getMinRow())
        # print (self.getMaxRow())
        # print (self.getMinColumn())
        # print (self.getMaxColumn())
        # print (self.getTargetRowObj(3))
        # print(self.getTargetRowObj(99))
        # print (self.getTargetRowObj(-3))
        # print (self.getTargetColumnObj(3))
        # print(self.getTargetColumnObj(22))
        # print(self.getTargetColumnObj(-3))
        # print (self.getTargetCellObj(2,3))
        # print (self.getTargetCellValue(3,5))
        # self.writeCellStrValue(2,7,"pass","green")
        # self.writeCellStrValue(3,7,"fail","red")
        # self.clearColumnValue("百度",6)
        # self.clearColumnValue("百度",7)
        # self.clearColumnValue("测试用例", 5)
        self.clearValue("测试用例", 2, 5)
        self.clearValue("测试用例", 2, 6)
        self.writeCellStrValue(2,6,getCurrentDateTime())

if __name__=="__main__":
    # excelObj = ParseExcel("sadf")
    excelObj=ParseExcel(testCaseFile)
    excelObj.printTestMethod()